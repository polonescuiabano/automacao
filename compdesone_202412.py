import openpyxl
from pymongo import MongoClient
from datetime import datetime

# Conexão com MongoDB
client = MongoClient(
    "mongodb://admin:3ng3nh4r1427611CPLUSPLUSOP@serorc.serpra.com.br:27017/serorc?authSource=admin"
)
db = client['serorc']
composicoes_collection = db['composicoes']


def processar_preco(valor):
    if valor is None:
        return None
    preco_str = str(valor).strip().replace(".", "").replace(",", ".")
    try:
        return float(preco_str)
    except ValueError:
        print(f"Valor inválido: {valor}")
        return None


def adicionar_precos_desonerados_dez2024(caminho_arquivo):
    wb = openpyxl.load_workbook(caminho_arquivo)
    sheet = wb.active

    data_cotacao = datetime(2024, 12, 1)
    tipo = "SINAPI"

    for row in range(7, 10000):
        descricao_classe = sheet.cell(row=row, column=1).value
        sigla_classe = sheet.cell(row=row, column=2).value
        codigo = sheet.cell(row=row, column=7).value
        descricao = sheet.cell(row=row, column=8).value
        unidade_medida = sheet.cell(row=row, column=9).value
        preco_desonerado = sheet.cell(row=row, column=11).value  # Coluna K

        if not codigo:
            continue

        try:
            codigo = str(codigo).strip()
        except ValueError:
            continue

        preco_desonerado = processar_preco(preco_desonerado)
        if preco_desonerado is None:
            continue

        composicao = composicoes_collection.find_one({"codigo": codigo})

        if composicao:
            atualizado = False
            for preco in composicao.get("precos_cotacao", []):
                if preco["data_cotacao"] == data_cotacao:
                    preco["preco_desonerado"] = preco_desonerado
                    atualizado = True
                    break

            if atualizado:
                composicoes_collection.update_one(
                    {"_id": composicao["_id"]},
                    {"$set": {"precos_cotacao": composicao["precos_cotacao"]}}
                )
                print(f"[UPDATE] Código {codigo} atualizado com preco_desonerado 12/2024.")
            else:
                composicoes_collection.update_one(
                    {"_id": composicao["_id"]},
                    {"$push": {
                        "precos_cotacao": {
                            "preco_desonerado": preco_desonerado,
                            "preco_nao_desonerado": None,
                            "data_cotacao": data_cotacao
                        }
                    }}
                )
                print(f"[ADD DATA] Nova cotação 12/2024 adicionada para código {codigo}.")
        else:
            if all([descricao_classe, sigla_classe, descricao, unidade_medida]):
                nova_composicao = {
                    "tipo": tipo,
                    "descricao_classe": descricao_classe,
                    "sigla_classe": sigla_classe,
                    "codigo": codigo,
                    "descricao": descricao,
                    "unidade_medida": unidade_medida,
                    "precos_cotacao": [
                        {
                            "preco_desonerado": preco_desonerado,
                            "preco_nao_desonerado": None,
                            "data_cotacao": data_cotacao
                        }
                    ],
                    "composicoes_auxiliares": [],
                    "insumos": []
                }

                result = composicoes_collection.insert_one(nova_composicao)
                print(f"[INSERT] Composição {codigo} criada com ID {result.inserted_id}")
            else:
                print(f"[ERRO] Dados incompletos para criar composição {codigo}")


# Caminho do Excel com preços desonerados de dezembro
caminho_excel_desonerado = r"C:\Users\Dell\Downloads\SINAPI_Custo_Ref_Composicoes_Sintetico_MT_202412_Desonerado.xlsx"

# Executar
adicionar_precos_desonerados_dez2024(caminho_excel_desonerado)

# Fechar conexão
client.close()
