import openpyxl
from pymongo import MongoClient
from datetime import datetime

# Conectar ao MongoDB
client = MongoClient(
    "mongodb://admin:3ng3nh4r1427611CPLUSPLUSOP@serorc.serpra.com.br:27017/serorc?authSource=admin"
)
db = client['serorc']
composicoes_collection = db['composicoes']


def processar_preco(valor):
    if valor is None:
        return None
    preco_str = str(valor).strip().replace(".", "").replace(",", ".")
    try:
        return float(preco_str)
    except ValueError:
        print(f"Valor inválido: {valor}")
        return None


def processar_composicoes_excel(caminho_arquivo):
    wb = openpyxl.load_workbook(caminho_arquivo)
    sheet = wb.active

    data_cotacao = datetime(2024, 12, 1)
    tipo = "SINAPI"

    for row in range(7, 10000):
        descricao_classe = sheet.cell(row=row, column=1).value
        sigla_classe = sheet.cell(row=row, column=2).value
        codigo = sheet.cell(row=row, column=7).value
        descricao = sheet.cell(row=row, column=8).value
        unidade_medida = sheet.cell(row=row, column=9).value
        preco_nao_desonerado = sheet.cell(row=row, column=11).value  # K: Preço não desonerado

        if codigo is not None:
            try:
                codigo = str(codigo).strip()
            except ValueError:
                continue

        preco_nao_desonerado = processar_preco(preco_nao_desonerado)

        if not all([descricao_classe, sigla_classe, codigo, descricao, unidade_medida, preco_nao_desonerado is not None]):
            continue

        composicao_existente = composicoes_collection.find_one({"codigo": codigo})

        if composicao_existente:
            atualizado = False
            for preco in composicao_existente.get("precos_cotacao", []):
                if preco["data_cotacao"] == data_cotacao:
                    preco["preco_nao_desonerado"] = preco_nao_desonerado
                    atualizado = True
                    break

            if atualizado:
                composicoes_collection.update_one(
                    {"_id": composicao_existente["_id"]},
                    {"$set": {"precos_cotacao": composicao_existente["precos_cotacao"]}}
                )
                print(f"[UPDATE] Código {codigo} atualizado com preco_nao_desonerado.")
            else:
                composicoes_collection.update_one(
                    {"_id": composicao_existente["_id"]},
                    {"$push": {
                        "precos_cotacao": {
                            "preco_desonerado": None,
                            "preco_nao_desonerado": preco_nao_desonerado,
                            "data_cotacao": data_cotacao
                        }
                    }}
                )
                print(f"[ADD DATA] Nova cotação adicionada para código {codigo}.")
        else:
            nova_composicao = {
                "tipo": tipo,
                "descricao_classe": descricao_classe,
                "sigla_classe": sigla_classe,
                "codigo": codigo,
                "descricao": descricao,
                "unidade_medida": unidade_medida,
                "precos_cotacao": [
                    {
                        "preco_desonerado": None,
                        "preco_nao_desonerado": preco_nao_desonerado,
                        "data_cotacao": data_cotacao
                    }
                ],
                "composicoes_auxiliares": [],
                "insumos": []
            }

            result = composicoes_collection.insert_one(nova_composicao)
            print(f"[INSERT] Composição {codigo} inserida com ID {result.inserted_id}")


# Caminho do Excel
caminho_arquivo = r"C:\Users\Dell\Downloads\SINAPI_Custo_Ref_Composicoes_Sintetico_MT_202412_NaoDesonerado.xlsx"

# Executar
processar_composicoes_excel(caminho_arquivo)

# Fechar conexão
client.close()
